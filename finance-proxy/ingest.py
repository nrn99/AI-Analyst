import csv
import hashlib
import io
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

import vertexai
from openpyxl import load_workbook
from pypdf import PdfReader
from vertexai.preview.generative_models import GenerativeModel

from categories import FIXED_CATEGORIES, UNCATEGORIZED

_CATEGORY_HINTS = {
    "Housing": ["rent", "mortgage", "lease", "hyra"],
    "Utilities": ["electric", "electricity", "water", "internet", "phone", "wifi"],
    "Groceries": ["grocery", "supermarket", "ica", "coop", "lidl", "willys"],
    "Dining": ["restaurant", "cafe", "coffee", "bar", "pizza", "burger"],
    "Transport": ["uber", "taxi", "bus", "metro", "train", "sl", "tram"],
    "Travel": ["hotel", "flight", "airbnb", "booking", "ryanair", "sas"],
    "Shopping": ["amazon", "ikea", "h&m", "zara", "shop"],
    "Subscriptions": ["netflix", "spotify", "subscription", "adobe"],
    "Health": ["pharmacy", "doctor", "clinic", "gym"],
    "Education": ["course", "tuition", "udemy", "coursera"],
    "Business": ["invoice", "client", "office", "supplies"],
    "Taxes": ["tax", "skatt"],
    "Fees": ["fee", "charge", "commission"],
    "Transfers": ["transfer", "bank", "swish"],
    "Savings/Investments": ["investment", "savings", "fund", "stock"],
    "Income": ["salary", "payroll", "income", "refund"],
}

_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%m-%d-%Y",
)

_MODEL = None
_MODEL_INITIALIZED = False


def _ensure_vertexai(project_id, location):
    global _MODEL_INITIALIZED
    if _MODEL_INITIALIZED:
        return
    vertexai.init(project=project_id, location=location)
    _MODEL_INITIALIZED = True


def _get_model(project_id, location):
    global _MODEL
    if _MODEL is None:
        _ensure_vertexai(project_id, location)
        _MODEL = GenerativeModel("gemini-2.0-flash")
    return _MODEL


def _normalize_text(value):
    return " ".join(str(value or "").strip().split())


def _normalize_merchant(value):
    lowered = _normalize_text(value).lower()
    lowered = re.sub(r"\d+", " ", lowered)
    lowered = re.sub(r"[^a-z\s]", " ", lowered)
    return " ".join(lowered.split())


def _normalize_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date().isoformat()
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _normalize_amount(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return str(value.quantize(Decimal("0.01")))
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(",", "")
    else:
        raw = raw.replace(",", ".")
    try:
        return str(Decimal(raw).quantize(Decimal("0.01")))
    except InvalidOperation:
        return None


def _detect_type(filename, content_type):
    if filename:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return "csv"
        if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
            return "xlsx"
        if lower.endswith(".pdf"):
            return "pdf"
    if content_type:
        if "csv" in content_type:
            return "csv"
        if "spreadsheet" in content_type:
            return "xlsx"
        if "pdf" in content_type:
            return "pdf"
    return "csv"


def _extract_metadata(text):
    metadata = {}
    if not text:
        return metadata
    patterns = {
        "account_holder": r"Account Holder:\s*(.+)",
        "account_type": r"Account Type:\s*(.+)",
        "account_number": r"Account Number:\s*(.+)",
        "reporting_period": r"Reporting Period:\s*(.+)",
        "currency": r"Currency:\s*([A-Z]{3})",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            metadata[key] = match.group(1).strip().strip(".")
    return metadata


def _header_map(headers):
    normalized = [str(h or "").strip().lower() for h in headers]
    mapping = {}
    header_aliases = {
        "date": ["date", "datum", "transaction date", "booked", "bokforingsdag"],
        "description": ["description", "beskrivning", "text", "merchant", "details"],
        "amount": ["amount", "belopp", "sum", "total"],
        "debit": ["debit", "withdrawal", "ut", "debet"],
        "credit": ["credit", "deposit", "in", "kredit"],
        "currency": ["currency", "valuta"],
    }
    for idx, header in enumerate(normalized):
        for key, aliases in header_aliases.items():
            if header in aliases:
                mapping[key] = idx
    return mapping


def _parse_csv(data):
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except Exception:
            continue
    if text is None:
        text = data.decode("utf-8", errors="ignore")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text.splitlines()[0])
    except Exception:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return [], {}

    header = rows[0]
    mapping = _header_map(header)
    transactions = []
    start_idx = 1 if mapping else 0
    for idx, row in enumerate(rows[start_idx:], start=start_idx):
        if mapping:
            date_val = row[mapping.get("date", 0)] if mapping.get("date") is not None else None
            desc_val = row[mapping.get("description", 1)] if mapping.get("description") is not None else None
            amount_val = None
            if "amount" in mapping:
                amount_val = row[mapping["amount"]]
            else:
                debit = row[mapping.get("debit")] if mapping.get("debit") is not None else None
                credit = row[mapping.get("credit")] if mapping.get("credit") is not None else None
                debit_norm = _normalize_amount(debit) or "0"
                credit_norm = _normalize_amount(credit) or "0"
                amount_val = str(Decimal(credit_norm) - Decimal(debit_norm))
            currency_val = row[mapping.get("currency")] if mapping.get("currency") is not None else None
        else:
            date_val = row[0] if len(row) > 0 else None
            desc_val = row[1] if len(row) > 1 else None
            amount_val = row[2] if len(row) > 2 else None
            currency_val = row[3] if len(row) > 3 else None

        transactions.append(
            {
                "date": date_val,
                "description": desc_val,
                "amount": amount_val,
                "currency": currency_val,
                "source_row": idx + 1,
            }
        )
    metadata = _extract_metadata(text)
    return transactions, metadata


def _parse_xlsx(data):
    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rows.append([cell for cell in row])

    if not rows:
        return [], {}

    mapping = _header_map(rows[0])
    transactions = []
    start_idx = 1 if mapping else 0
    for idx, row in enumerate(rows[start_idx:], start=start_idx):
        row = list(row)
        if mapping:
            date_val = row[mapping.get("date", 0)] if mapping.get("date") is not None else None
            desc_val = row[mapping.get("description", 1)] if mapping.get("description") is not None else None
            amount_val = None
            if "amount" in mapping:
                amount_val = row[mapping["amount"]]
            else:
                debit = row[mapping.get("debit")] if mapping.get("debit") is not None else None
                credit = row[mapping.get("credit")] if mapping.get("credit") is not None else None
                debit_norm = _normalize_amount(debit) or "0"
                credit_norm = _normalize_amount(credit) or "0"
                amount_val = str(Decimal(credit_norm) - Decimal(debit_norm))
            currency_val = row[mapping.get("currency")] if mapping.get("currency") is not None else None
        else:
            date_val = row[0] if len(row) > 0 else None
            desc_val = row[1] if len(row) > 1 else None
            amount_val = row[2] if len(row) > 2 else None
            currency_val = row[3] if len(row) > 3 else None

        transactions.append(
            {
                "date": date_val,
                "description": desc_val,
                "amount": amount_val,
                "currency": currency_val,
                "source_row": idx + 1,
            }
        )

    text_sample = "\n".join(str(cell) for cell in rows[0][:6] if cell)
    metadata = _extract_metadata(text_sample)
    return transactions, metadata


def _parse_pdf(data):
    reader = PdfReader(io.BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    metadata = _extract_metadata(text)
    transactions = []
    for idx, line in enumerate(text.splitlines()):
        line = line.strip()
        if not line:
            continue
        date_match = re.search(r"\b(\d{4}-\d{2}-\d{2}|\d{2}[./-]\d{2}[./-]\d{4})\b", line)
        amount_matches = re.findall(r"[-+]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?", line)
        if not date_match or not amount_matches:
            continue
        date_val = date_match.group(1)
        amount_val = amount_matches[-1]
        desc = line.replace(date_val, "").replace(amount_val, "").strip(" -")
        transactions.append(
            {
                "date": date_val,
                "description": desc,
                "amount": amount_val,
                "currency": metadata.get("currency"),
                "source_row": idx + 1,
            }
        )
    return transactions, metadata


def _suggest_category_heuristic(description, amount):
    text = _normalize_text(description).lower()
    for category, keywords in _CATEGORY_HINTS.items():
        for keyword in keywords:
            if keyword in text:
                return category
    if amount is not None:
        try:
            if Decimal(str(amount)) > 0:
                return "Income"
        except Exception:
            pass
    return UNCATEGORIZED


def _parse_model_response(response_text):
    if not response_text:
        return None
    normalized = response_text.strip().strip(".").strip()
    for category in FIXED_CATEGORIES:
        if normalized.lower() == category.lower():
            return category
    return None


def suggest_category(description, amount, project_id, location):
    mode = os.getenv("CATEGORY_SUGGESTION_MODE", "model").lower()
    if mode == "model":
        try:
            model = _get_model(project_id, location)
            prompt = (
                "Choose exactly one category from this list:\\n"
                + ", ".join(FIXED_CATEGORIES)
                + "\\n"
                + f"Transaction: description={description!r}, amount={amount!r}. "
                "Respond with only the category name."
            )
            response = model.generate_content(prompt)
            category = _parse_model_response(getattr(response, "text", ""))
            if category:
                return category
        except Exception:
            pass
    return _suggest_category_heuristic(description, amount)


def parse_statement(data, filename, content_type, project_id, location):
    file_type = _detect_type(filename, content_type)
    if file_type == "xlsx":
        raw_transactions, metadata = _parse_xlsx(data)
    elif file_type == "pdf":
        raw_transactions, metadata = _parse_pdf(data)
    else:
        raw_transactions, metadata = _parse_csv(data)

    file_hash = hashlib.sha256(data).hexdigest()
    batch_id = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{file_hash[:8]}"
    currency = metadata.get("currency")

    normalized = []
    for row in raw_transactions:
        date_value = _normalize_date(row.get("date"))
        amount_value = _normalize_amount(row.get("amount"))
        description = _normalize_text(row.get("description")) or "Unknown"
        if not date_value or amount_value is None:
            continue
        merchant_raw = description
        merchant_normalized = _normalize_merchant(description)
        category = suggest_category(description, amount_value, project_id, location)
        normalized.append(
            {
                "date": date_value,
                "description": description,
                "amount": amount_value,
                "currency": row.get("currency") or currency,
                "merchant_raw": merchant_raw,
                "merchant_normalized": merchant_normalized,
                "category_suggested": category,
                "needs_review": category == UNCATEGORIZED,
                "source_row": row.get("source_row"),
            }
        )

    return {
        "batch_id": batch_id,
        "file_hash": file_hash,
        "metadata": metadata,
        "transactions": normalized,
    }
