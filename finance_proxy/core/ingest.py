import csv
import hashlib
import io
import logging
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

import vertexai
from openpyxl import load_workbook
from pypdf import PdfReader
from vertexai.preview.generative_models import GenerativeModel

from .categories import FIXED_CATEGORIES, UNCATEGORIZED

_CATEGORY_HINTS = {
    "Housing": ["rent", "mortgage", "lease", "hyra"],
    "Utilities": ["electric", "electricity", "water", "internet", "phone", "wifi"],
    "Groceries": ["grocery", "supermarket", "ica", "coop", "lidl", "willys"],
    "Dining": ["restaurant", "cafe", "coffee", "bar", "pizza", "burger"],
    "Transport": ["uber", "taxi", "bus", "metro", "train", "sl", "tram"],
    "Travel": ["hotel", "flight", "airbnb", "booking", "ryanair", "sas"],
    "Shopping": ["amazon", "ikea", "h&m", "zara", "shop"],
    "Subscriptions": ["netflix", "spotify", "subscription", "adobe"],
    "Health": ["pharmacy", "doctor", "clinic", "gym"],
    "Education": ["course", "tuition", "udemy", "coursera"],
    "Business": ["invoice", "client", "office", "supplies"],
    "Taxes": ["tax", "skatt"],
    "Fees": ["fee", "charge", "commission"],
    "Tithe": ["church", "tithe", "tionde", "we are one church", "hillsong", "filadelfia"],
    "Charity": ["charity", "donation", "gift", "red cross", "unicef"],
    "Overföring": ["överföring", "overföring", "internal transfer", "balance movement", "account transfer", "egen överföring"],
    "Transfers": ["transfer", "bank", "swish"],
    "Savings/Investments": ["investment", "savings", "fund", "stock"],
    "Income": ["salary", "payroll", "income", "refund"],
}

_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%m-%d-%Y",
)

_MODEL = None
_MODEL_INITIALIZED = False


def _ensure_vertexai(project_id, location):
    global _MODEL_INITIALIZED
    if _MODEL_INITIALIZED:
        return
    vertexai.init(project=project_id, location=location)
    _MODEL_INITIALIZED = True


def _get_model(project_id, location):
    global _MODEL
    if _MODEL is None:
        _ensure_vertexai(project_id, location)
        _MODEL = GenerativeModel("gemini-2.0-flash")
    return _MODEL


def _normalize_text(value):
    return " ".join(str(value or "").strip().split())


def _normalize_merchant(value):
    lowered = _normalize_text(value).lower()
    lowered = re.sub(r"\d+", " ", lowered)
    lowered = re.sub(r"[^a-z\s]", " ", lowered)
    return " ".join(lowered.split())


def _normalize_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date().isoformat()
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _normalize_amount(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return str(value.quantize(Decimal("0.01")))
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(",", "")
    else:
        raw = raw.replace(",", ".")
    try:
        return str(Decimal(raw).quantize(Decimal("0.01")))
    except InvalidOperation:
        return None


def _detect_type(filename, content_type):
    if filename:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return "csv"
        if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
            return "xlsx"
        if lower.endswith(".pdf"):
            return "pdf"
    if content_type:
        if "csv" in content_type:
            return "csv"
        if "spreadsheet" in content_type:
            return "xlsx"
        if "pdf" in content_type:
            return "pdf"
    return "csv"


def _extract_metadata(text):
    metadata = {}
    if not text:
        return metadata
    patterns = {
        "account_holder": r"Account Holder:\s*(.+)",
        "account_type": r"Account Type:\s*(.+)",
        "account_number": r"Account Number:\s*(.+)",
        "reporting_period": r"Reporting Period:\s*(.+)",
        "currency": r"Currency:\s*([A-Z]{3})",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            metadata[key] = match.group(1).strip().strip(".")
    return metadata


def _header_map(headers):
    normalized = [str(h or "").strip().lower() for h in headers]
    mapping = {}
    # Restricted Swedbank Mapping
    header_aliases = {
        "date": ["bokföringsdag", "transaktionsdag"],
        "description": ["beskrivning"],
        "amount": ["belopp"],
        "currency": ["valuta"],
        # Optional metadata fields
        "reference": ["referens"],
        "balance": ["bokfört saldo"],
        "clearing": ["clearingnummer"],
        "account": ["kontonummer"],
        "product": ["produkt"],
    }
    for idx, header in enumerate(normalized):
        for key, aliases in header_aliases.items():
            if header in aliases:
                mapping[key] = idx
    return mapping


def _find_header(rows, limit=20):
    """
    Scans the first `limit` rows to find the Swedbank header row.
    """
    for idx, row in enumerate(rows[:limit]):
        headers = [str(cell) for cell in row if cell is not None]
        mapping = _header_map(headers)
        
        # Strict check for critical Swedbank fields
        has_critical = "date" in mapping and "amount" in mapping and "description" in mapping
        
        if has_critical:
            LOGGER = logging.getLogger("finance_proxy")
            LOGGER.info(f"Swedbank header found at row {idx}: {row}")
            return idx, mapping

    LOGGER = logging.getLogger("finance_proxy")
    LOGGER.warning("No valid Swedbank header row found.")
    return 0, {}


def _parse_csv(data):
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except Exception:
            continue
    if text is None:
        text = data.decode("utf-8", errors="ignore")

    sniffer = csv.Sniffer()
    dialect = None
    try:
        # Give sniffer a hint to prefer common delimiters
        dialect = sniffer.sniff(text.splitlines()[0], delimiters=[",", ";", "\t"])
    except Exception:
        pass
    
    # Fallback manual detection if sniffer fails
    if not dialect:
        first_line = text.splitlines()[0]
        if "," in first_line:
            dialect = csv.excel
        elif ";" in first_line:
            class SemiColonDialect(csv.Dialect):
                delimiter = ';'
                quotechar = '"'
                doublequote = True
                skipinitialspace = False
                lineterminator = '\r\n'
                quoting = csv.QUOTE_MINIMAL
            dialect = SemiColonDialect
        else:
            dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return [], {}

    header_idx, mapping = _find_header(rows)
    transactions = []
    
    # Start usually after the header
    start_idx = header_idx + 1

    for idx, row in enumerate(rows[start_idx:], start=start_idx):
        if mapping:
            date_val = row[mapping.get("date")] if "date" in mapping and mapping["date"] < len(row) else None
            desc_val = row[mapping.get("description")] if "description" in mapping and mapping["description"] < len(row) else None
            
            amount_val = None
            if "amount" in mapping and mapping["amount"] < len(row):
                amount_val = row[mapping["amount"]]
            else:
                debit = row[mapping.get("debit")] if "debit" in mapping and mapping["debit"] < len(row) else None
                credit = row[mapping.get("credit")] if "credit" in mapping and mapping["credit"] < len(row) else None
                debit_norm = _normalize_amount(debit) or "0"
                credit_norm = _normalize_amount(credit) or "0"
                if debit or credit:
                    amount_val = str(Decimal(credit_norm) - Decimal(debit_norm))
            
            currency_val = row[mapping.get("currency")] if "currency" in mapping and mapping["currency"] < len(row) else None
        else:
            # Fallback 0,1,2,3
            date_val = row[0] if len(row) > 0 else None
            desc_val = row[1] if len(row) > 1 else None
            amount_val = row[2] if len(row) > 2 else None
            currency_val = row[3] if len(row) > 3 else None

        transactions.append(
            {
                "date": date_val,
                "description": desc_val,
                "amount": amount_val,
                "currency": currency_val,
                "source_row": idx + 1,
            }
        )
    metadata = _extract_metadata(text)
    return transactions, metadata


def _parse_xlsx(data):
    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rows.append([cell for cell in row])

    if not rows:
        return [], {}

    header_idx, mapping = _find_header(rows)
    transactions = []
    start_idx = header_idx + 1

    for idx, row in enumerate(rows[start_idx:], start=start_idx):
        row = list(row)
        if mapping:
            date_val = row[mapping.get("date")] if "date" in mapping and mapping["date"] < len(row) else None
            desc_val = row[mapping.get("description")] if "description" in mapping and mapping["description"] < len(row) else None
            
            amount_val = None
            if "amount" in mapping and mapping["amount"] < len(row):
                amount_val = row[mapping["amount"]]
            else:
                debit = row[mapping.get("debit")] if "debit" in mapping and mapping["debit"] < len(row) else None
                credit = row[mapping.get("credit")] if "credit" in mapping and mapping["credit"] < len(row) else None
                debit_norm = _normalize_amount(debit) or "0"
                credit_norm = _normalize_amount(credit) or "0"
                if debit or credit:
                    amount_val = str(Decimal(credit_norm) - Decimal(debit_norm))
            
            currency_val = row[mapping.get("currency")] if "currency" in mapping and mapping["currency"] < len(row) else None
        else:
            date_val = row[0] if len(row) > 0 else None
            desc_val = row[1] if len(row) > 1 else None
            amount_val = row[2] if len(row) > 2 else None
            currency_val = row[3] if len(row) > 3 else None

        transactions.append(
            {
                "date": date_val,
                "description": desc_val,
                "amount": amount_val,
                "currency": currency_val,
                "source_row": idx + 1,
            }
        )

    text_sample = "\n".join(str(cell) for cell in rows[0][:6] if cell)
    metadata = _extract_metadata(text_sample)
    return transactions, metadata


def _parse_pdf(data):
    reader = PdfReader(io.BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    metadata = _extract_metadata(text)
    transactions = []
    for idx, line in enumerate(text.splitlines()):
        line = line.strip()
        if not line:
            continue
        date_match = re.search(r"\b(\d{4}-\d{2}-\d{2}|\d{2}[./-]\d{2}[./-]\d{4})\b", line)
        amount_matches = re.findall(r"[-+]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?", line)
        if not date_match or not amount_matches:
            continue
        date_val = date_match.group(1)
        amount_val = amount_matches[-1]
        desc = line.replace(date_val, "").replace(amount_val, "").strip(" -")
        transactions.append(
            {
                "date": date_val,
                "description": desc,
                "amount": amount_val,
                "currency": metadata.get("currency"),
                "source_row": idx + 1,
            }
        )
    return transactions, metadata


def _suggest_category_heuristic(description, amount):
    text = _normalize_text(description).lower()
    for category, keywords in _CATEGORY_HINTS.items():
        for keyword in keywords:
            if keyword in text:
                return category
    if amount is not None:
        try:
            if Decimal(str(amount)) > 0:
                return "Income"
        except Exception:
            pass
    return UNCATEGORIZED


def _parse_model_response(response_text):
    if not response_text:
        return None
    normalized = response_text.strip().strip(".").strip()
    for category in FIXED_CATEGORIES:
        if normalized.lower() == category.lower():
            return category
    return None


def suggest_category(description, amount, project_id, location):
    mode = os.getenv("CATEGORY_SUGGESTION_MODE", "model").lower()
    if mode == "model":
        try:
            model = _get_model(project_id, location)
            prompt = (
                "Choose exactly one category from this list:\\n"
                + ", ".join(FIXED_CATEGORIES)
                + "\\n"
                + f"Transaction: description={description!r}, amount={amount!r}. "
                "Respond with only the category name."
            )
            response = model.generate_content(prompt)
            category = _parse_model_response(getattr(response, "text", ""))
            if category:
                return category
        except Exception:
            pass
    return _suggest_category_heuristic(description, amount)


def _derive_pillar(category):
    category = (category or "").lower().strip()
    if category in ["overföring", "transfers"]:
        return "Internal"
    if category in ["rent", "housing", "utilities", "groceries", "transport", "health"]:
        return "Needs"
    if category in ["dining", "shopping", "subscriptions", "travel"]:
        return "Wants"
    if category in ["tithe", "charity"]:
        return "Faith"
    if category in ["savings/investments", "education", "business"]:
        return "Growth"
    return ""


def parse_statement(data, filename, content_type, project_id, location):
    file_type = _detect_type(filename, content_type)
    LOGGER = logging.getLogger("finance_proxy")
    LOGGER.info(f"Parsing file: {filename} ({content_type}), type detected: {file_type}")
    
    if file_type != "csv":
        raise ValueError("Only CSV files are supported for now (Swedbank format)")
        
    raw_transactions, metadata = _parse_csv(data)

    LOGGER.info(f"Raw transactions extracted: {len(raw_transactions)}")
    if raw_transactions:
        LOGGER.info(f"Sample raw row: {raw_transactions[0]}")

    file_hash = hashlib.sha256(data).hexdigest()
    batch_id = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{file_hash[:8]}"
    currency = metadata.get("currency")

    normalized = []
    for row in raw_transactions:
        date_value = _normalize_date(row.get("date"))
        amount_value = _normalize_amount(row.get("amount"))
        description = _normalize_text(row.get("description")) or "Unknown"
        
        if not date_value:
            LOGGER.warning(f"Dropping row due to invalid date: {row.get('date')} -> {date_value}")
            continue
        if amount_value is None:
            LOGGER.warning(f"Dropping row due to invalid amount: {row.get('amount')} -> {amount_value}")
            continue
            
        merchant_raw = description
        merchant_normalized = _normalize_merchant(description)
        # category = suggest_category(description, amount_value, project_id, location)
        # Optimization: use heuristic first to avoid API spam during debugging
        category = _suggest_category_heuristic(description, amount_value)
        machine_pillar = _derive_pillar(category)
        
        normalized.append(
            {
                "date": date_value,
                "description": description,
                "amount": amount_value,
                "currency": row.get("currency") or currency,
                "merchant_raw": merchant_raw,
                "merchant_normalized": merchant_normalized,
                "category_suggested": category,
                "machine_pillar": machine_pillar,
                "integrity_filter": "Planned",
                "root_trigger": "",
                "notes": "",
                "needs_review": category == UNCATEGORIZED,
                "source_row": row.get("source_row"),
            }
        )
    
    LOGGER.info(f"Normalized transactions: {len(normalized)}")

    return {
        "batch_id": batch_id,
        "file_hash": file_hash,
        "metadata": metadata,
        "transactions": normalized,
    }
